from datetime import date, datetime
import io
import os
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.utils import timezone
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.utils.timezone import now, localdate
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from django.templatetags.static import static
from django.db.models import Q, Prefetch
import csv
from io import BytesIO
from django.urls import reverse
from django.http import HttpResponseRedirect

from .models import AccessRequest, RequestedSystem, UserRole
from .forms import AccessRequestForm

# --- HELPER: Centralized Status Logic ---
def sync_request_status(request_obj):
    all_systems = request_obj.requested_systems.all()
    
    hod_pending = all_systems.filter(hod_status='pending').exists()
    if hod_pending:
        request_obj.status = 'pending_hod'
        request_obj.save()
        return

    hod_approved_exists = all_systems.filter(hod_status='approved').exists()
    if not hod_approved_exists:
        request_obj.status = 'rejected_hod'
        request_obj.save()
        return

    ict_pending = all_systems.filter(hod_status='approved', ict_status='pending').exists()
    if ict_pending:
        request_obj.status = 'pending_ict'
        request_obj.save()
        return

    ict_approved_exists = all_systems.filter(ict_status='approved').exists()
    if ict_approved_exists:
        request_obj.status = 'approved'
    else:
        request_obj.status = 'rejected_ict'
    
    request_obj.save()

# --- VIEWS ---

@login_required
def request_form_view(request):
    if request.method == 'POST':
        form = AccessRequestForm(request.POST)
        if form.is_valid():
            access = form.save(commit=False)
            access.requester = request.user
            access.tsc_no = request.user.tsc_no
            access.email = request.user.email
            access.directorate = request.user.directorate
            access.save()
        
            for system in form.cleaned_data['systems']:
                RequestedSystem.objects.create(
                    access_request=access,
                    system=system,
                    level_of_access=form.cleaned_data.get('access_levels', 'User')
                )

            if access.directorate and access.directorate.hod_email:
                send_mail(
                    subject='[TSC] New System Access Request Awaiting Your Approval',
                    message=f"A new access request from {request.user.get_full_name()} ({request.user.email}) is pending review.",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[access.directorate.hod_email],
                )

            send_mail(
                subject='[TSC] Your System Access Request Has Been Submitted',
                message=f"Hi {request.user.get_full_name()},\n\nYour request has been submitted and sent to your HOD.",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[request.user.email],
            )

            return redirect('request_submitted')
    else:
        form = AccessRequestForm(initial={
            'tsc_no': request.user.tsc_no,
            'email': request.user.email,
            'directorate': request.user.directorate
        })
    return render(request, 'access_request/request_form.html', {'form': form})

@login_required
def submit_request(request):
    return request_form_view(request)

@login_required
def hod_dashboard(request):
    """HOD dashboard: Pending items + History + Search/Filter + Export."""
    user = request.user
    directorate = None

    # 1. Role Guard - User must be HOD
    user_role = UserRole.objects.filter(user=user, role='hod').first()
    if not user_role:
        messages.error(request, "You do not have access to HOD dashboard.")
        return redirect('user_home')
    
    # 2. Get Directorate from UserRole
    directorate = user_role.directorate
    if not directorate:
        messages.error(request, "No directorate assignment found for your HOD role.")
        return redirect('user_home')

    requests = AccessRequest.objects.none()
    history = AccessRequest.objects.none()
    
    # 2. Capture Filter Parameters
    search_term = request.GET.get('tsc', "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    active_tab = request.GET.get("active_tab", "pending")

    if directorate:
        # --- A. Pending Requests ---
        # Fetch AccessRequests (Parents) that have at least one pending system for this directorate
        requests = AccessRequest.objects.filter(
            directorate=directorate,
            requested_systems__hod_status="pending"
        ).distinct().select_related('requester').prefetch_related(
            Prefetch('requested_systems', queryset=RequestedSystem.objects.filter(hod_status='pending'))
        )

        # --- B. History Requests ---
        # Requests where the current user is recorded as the hod_approver
        history = AccessRequest.objects.filter(
            hod_approver=user
        ).order_by('-submitted_at').select_related('requester').prefetch_related('requested_systems')

        # --- C. Apply TSC Search (Fixes your issue) ---
        if search_term:
            requests = requests.filter(tsc_no__icontains=search_term)
            history = history.filter(tsc_no__icontains=search_term)

        # --- D. Apply Date Range Filter ---
        if start_date and end_date:
            try:
                s_date = datetime.strptime(start_date, "%Y-%m-%d")
                # Set end date to end of day (23:59:59)
                e_date = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                
                requests = requests.filter(submitted_at__range=(s_date, e_date))
                history = history.filter(submitted_at__range=(s_date, e_date))
            except ValueError:
                pass

        # --- E. Export Logic (Exports HISTORY data) ---
        if "export_excel" in request.GET:
            wb = Workbook()
            ws = wb.active
            ws.title = "HOD Decisions"
            ws.append(["Requester", "TSC No", "Designation", "System", "Decision", "Action Date", "Comment"])
            
            for req in history:
                for sys in req.requested_systems.all():
                    if sys.hod_status != 'pending':
                        action_date = sys.hod_decision_date.strftime("%Y-%m-%d %H:%M") if sys.hod_decision_date else "-"
                        ws.append([
                            req.requester.full_name, req.tsc_no, req.designation,
                            sys.get_system_display(), sys.hod_status.upper(), 
                            action_date, sys.hod_comment
                        ])
            
            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = f'attachment; filename="HOD_Report_{localdate()}.xlsx"'
            wb.save(response)
            return response

        if "export_pdf" in request.GET:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()
            
            elements.append(Paragraph(f"HOD Approval Report - {directorate.name}", styles["Title"]))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d')}", styles["Normal"]))
            elements.append(Spacer(1, 20))
            
            data = [["Requester", "TSC No", "System", "Decision", "Date"]]
            for req in history:
                for sys in req.requested_systems.all():
                    if sys.hod_status != 'pending':
                        d_date = sys.hod_decision_date.strftime("%Y-%m-%d") if sys.hod_decision_date else "-"
                        data.append([
                            req.requester.full_name, req.tsc_no, 
                            sys.get_system_display(), sys.hod_status.upper(), d_date
                        ])

            table = Table(data, colWidths=[140, 80, 140, 80, 80])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
            ]))
            elements.append(table)
            doc.build(elements)
            
            buffer.seek(0)
            return HttpResponse(buffer, content_type='application/pdf')

    else:
        messages.error(request, "No directorate assignment found.")

    return render(request, "access_request/hod_dashboard.html", {
        "requests": requests,
        "history": history,
        "hod_directorate": directorate,
        "user": user,
        "active_tab": active_tab,
    })

    
@login_required
def hod_system_decision(request, system_id):
    # ... (Role & Scope checks remain the same) ...
    if not getattr(request.user, "userrole", None) or request.user.userrole.role != "hod":
        return HttpResponse(status=403, content="Access Denied")

    system = get_object_or_404(RequestedSystem, id=system_id)
    request_obj = system.access_request
    
    # Check if user is HOD for this directorate or manages the requester
    hod_role = UserRole.objects.filter(user=request.user, role='hod').first()
    is_authorized = hod_role and (
        hod_role.directorate == request_obj.directorate or 
        UserRole.objects.filter(hod=request.user, user=request_obj.requester).exists()
    )
    if not is_authorized:
         return HttpResponse(status=403, content="Access Denied")

    if request.method == "POST":
        action = request.POST.get("action")
        comment = request.POST.get("comment", "")
        sys_name = system.get_system_display()

        system.hod_decision_date = timezone.now()
        request_obj.hod_approver = request.user

        if action == "approve":
            system.hod_status = "approved"
            system.hod_comment = ""
            system.save()
            messages.success(request, f"✅ Approved {sys_name}.")
        elif action == "reject":
            system.hod_status = "rejected"
            system.hod_comment = comment
            system.ict_status = "rejected"
            system.sysadmin_status = "rejected"
            system.save()
            messages.warning(request, f"❌ Rejected {sys_name}.")

        sync_request_status(request_obj)
        request_obj.save()

        # Email Logic (Keep existing)
        pending = request_obj.requested_systems.filter(hod_status="pending").exists()
        if not pending:
            # ... (Keep existing email code) ...
            pass

        # ✅ FIX: Redirect with Preserved Filters
        base_url = reverse('hod_dashboard')
        query_params = request.GET.urlencode() # Captures 'tsc=...', 'active_tab=...', etc.
        if query_params:
            return HttpResponseRedirect(f"{base_url}?{query_params}")
        return redirect("hod_dashboard")

    return redirect("hod_dashboard")

@login_required
def ict_dashboard(request):
    """ICT dashboard: Pending items + History + Search/Filter + Export."""
    user = request.user
    
    # 1. Role Guard - User must be ICT
    user_role = UserRole.objects.filter(user=user, role='ict').first()
    if not user_role:
        messages.error(request, "You do not have access to ICT dashboard.")
        return redirect('user_home')

    # 2. Base Querysets
    requests = AccessRequest.objects.none()
    history = AccessRequest.objects.none()

    # 2. Capture Filter Parameters
    search_term = request.GET.get('tsc', "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    active_tab = request.GET.get("active_tab", "pending")

    # --- A. Pending Requests ---
    # Fetch AccessRequests (Parents) that have systems Approved by HOD but Pending ICT
    requests = AccessRequest.objects.filter(
        requested_systems__hod_status="approved",
        requested_systems__ict_status="pending"
    ).distinct().select_related('requester', 'directorate').prefetch_related(
        Prefetch('requested_systems', queryset=RequestedSystem.objects.filter(hod_status='approved', ict_status='pending'))
    )

    # --- B. History Requests ---
    # Requests where the current user is recorded as the ict_approver
    history = AccessRequest.objects.filter(
        ict_approver=user
    ).order_by('-submitted_at').select_related('requester', 'directorate').prefetch_related('requested_systems')

    # --- C. Apply TSC Search ---
    if search_term:
        requests = requests.filter(tsc_no__icontains=search_term)
        history = history.filter(tsc_no__icontains=search_term)

    # --- D. Apply Date Range Filter ---
    if start_date and end_date:
        try:
            s_date = datetime.strptime(start_date, "%Y-%m-%d")
            e_date = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            
            requests = requests.filter(submitted_at__range=(s_date, e_date))
            history = history.filter(submitted_at__range=(s_date, e_date))
        except ValueError:
            pass

    # --- E. Export Logic (Exports HISTORY data) ---
    if "export_excel" in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.title = "ICT Decisions"
        ws.append(["Requester", "TSC No", "Directorate", "System", "Decision", "Action Date", "Comment"])
        
        for req in history:
            for sys in req.requested_systems.all():
                # Export only items actioned by ICT (not pending)
                if sys.ict_status != 'pending':
                    action_date = sys.ict_decision_date.strftime("%Y-%m-%d %H:%M") if sys.ict_decision_date else "-"
                    ws.append([
                        req.requester.full_name, req.tsc_no, 
                        req.directorate.name if req.directorate else "-",
                        sys.get_system_display(), sys.ict_status.upper(), 
                        action_date, sys.ict_comment
                    ])
        
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="ICT_Report_{localdate()}.xlsx"'
        wb.save(response)
        return response

    if "export_pdf" in request.GET:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph(f"ICT Approval Report", styles["Title"]))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d')}", styles["Normal"]))
        elements.append(Spacer(1, 20))
        
        data = [["Requester", "TSC No", "System", "Decision", "Date"]]
        for req in history:
            for sys in req.requested_systems.all():
                if sys.ict_status != 'pending':
                    d_date = sys.ict_decision_date.strftime("%Y-%m-%d") if sys.ict_decision_date else "-"
                    data.append([
                        req.requester.full_name, req.tsc_no, 
                        sys.get_system_display(), sys.ict_status.upper(), d_date
                    ])

        table = Table(data, colWidths=[140, 80, 140, 80, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')

    return render(request, "access_request/ict_dashboard.html", {
        "requests": requests,
        "history": history,
        "user": user,
        "active_tab": active_tab,
    })


@login_required
def ict_system_decision(request, system_id):
    # ... (Role checks remain the same) ...
    if not getattr(request.user, "userrole", None) or request.user.userrole.role != "ict":
        return HttpResponse(status=403)

    system = get_object_or_404(RequestedSystem, id=system_id)
    request_obj = system.access_request

    if request.method == "POST":
        action = request.POST.get("action")
        comment = request.POST.get("comment", "")
        
        system.ict_decision_date = timezone.now()
        request_obj.ict_approver = request.user

        if action == "approve":
            system.ict_status = "approved"
            system.ict_comment = ""
            system.save()
            messages.success(request, f"{system.get_system_display()} approved.")
        elif action == "reject":
            system.ict_status = "rejected"
            system.ict_comment = comment
            system.save()
            messages.warning(request, f"{system.get_system_display()} rejected.")

        sync_request_status(request_obj)
        request_obj.save()

        # Email Logic (Keep existing)
        # ...

        # ✅ FIX: Redirect with Preserved Filters
        base_url = reverse('ict_dashboard')
        query_params = request.GET.urlencode()
        if query_params:
            return HttpResponseRedirect(f"{base_url}?{query_params}")
        return redirect("ict_dashboard")

    return redirect("ict_dashboard")



# access_request/views.py
# ... (Keep existing imports) ...

@login_required
def system_admin_dashboard(request):
    """System Admin Dashboard: Pending + History + Filters + Export."""
    # 1. Check if user is a System Admin
    user_role = UserRole.objects.filter(user=request.user, role='sys_admin').first()
    if not user_role:
        messages.error(request, "You do not have access to System Admin dashboard.")
        return redirect("user_home")

    # 2. Get the system assigned to this admin
    assigned_system = user_role.system_assigned
    if not assigned_system:
        messages.error(request, "No system has been assigned to you yet.")
        return redirect("user_home")
    
    # 3. Filter Parameters
    search_term = request.GET.get('tsc', "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    active_tab = request.GET.get("active_tab", "pending")
    
    # 4. Base Querysets - Filter ONLY by assigned system
    
    # A. Pending: Systems with status 'pending' for THIS admin's assigned system
    requests = AccessRequest.objects.filter(
        requested_systems__system=assigned_system,
        requested_systems__sysadmin_status="pending"
    ).distinct().select_related('requester').prefetch_related(
        Prefetch('requested_systems', queryset=RequestedSystem.objects.filter(
            system=assigned_system,
            sysadmin_status='pending'
        ))
    )

    # B. History: Systems actioned by THIS admin for THIS assigned system
    history = AccessRequest.objects.filter(
        requested_systems__system=assigned_system,
        requested_systems__system_admin=request.user
    ).distinct().order_by('-submitted_at').select_related('requester').prefetch_related(
        Prefetch('requested_systems', queryset=RequestedSystem.objects.filter(
            system=assigned_system,
            system_admin=request.user
        ))
    )

    # 5. Apply Filters
    if search_term:
        requests = requests.filter(tsc_no__icontains=search_term)
        history = history.filter(tsc_no__icontains=search_term)

    if start_date and end_date:
        try:
            s_date = datetime.strptime(start_date, "%Y-%m-%d")
            e_date = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            requests = requests.filter(submitted_at__range=(s_date, e_date))
            history = history.filter(submitted_at__range=(s_date, e_date))
        except ValueError:
            pass

    # 5. Export Logic (History)
    if "export_excel" in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.title = "System Admin History"
        ws.append(["Requester", "TSC No", "System", "Level", "Decision", "Action Date", "Comment"])
        
        for req in history:
            for sys in req.requested_systems.all():
                if sys.system_admin == request.user:
                    action_date = sys.sysadmin_decision_date.strftime("%Y-%m-%d %H:%M") if sys.sysadmin_decision_date else "-"
                    ws.append([
                        req.requester.full_name, req.tsc_no,
                        sys.get_system_display(), sys.level_of_access,
                        sys.sysadmin_status.upper(), action_date, sys.sysadmin_comment
                    ])
        
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="System_Admin_History_{localdate()}.xlsx"'
        wb.save(response)
        return response

    if "export_pdf" in request.GET:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph("System Admin - Approval History", styles["Title"]))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d')}", styles["Normal"]))
        elements.append(Spacer(1, 20))
        
        data = [["Requester", "TSC No", "System", "Level", "Decision", "Date"]]
        for req in history:
            for sys in req.requested_systems.all():
                if sys.system_admin == request.user:
                    d_date = sys.sysadmin_decision_date.strftime("%Y-%m-%d") if sys.sysadmin_decision_date else "-"
                    data.append([
                        req.requester.full_name, req.tsc_no,
                        sys.get_system_display(), sys.level_of_access, sys.sysadmin_status.upper(), d_date
                    ])

        table = Table(data, colWidths=[140, 80, 100, 80, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')

    # Stats Counters - Only for assigned system
    total_requests = RequestedSystem.objects.filter(system=assigned_system, sysadmin_status__isnull=False).count()
    pending_requests = RequestedSystem.objects.filter(system=assigned_system, sysadmin_status="pending").count()
    approved_requests = RequestedSystem.objects.filter(system=assigned_system, sysadmin_status="approved").count()
    rejected_requests = RequestedSystem.objects.filter(system=assigned_system, sysadmin_status="rejected").count()
    today_requests = RequestedSystem.objects.filter(system=assigned_system, access_request__submitted_at__date=date.today()).count()

    # Get system name
    system_name = dict(RequestedSystem.SYSTEM_CHOICES).get(assigned_system, assigned_system)

    context = {
        "system_name": system_name,
        "requests": requests,
        "history": history,
        "total_requests": total_requests,
        "pending_requests": pending_requests,
        "approved_requests": approved_requests,
        "rejected_requests": rejected_requests,
        "today_requests": today_requests,
        "active_tab": active_tab,
    }
    return render(request, "access_request/system_admin_dashboard.html", context)



@require_POST
@login_required
def system_admin_decision(request, pk):
    # 1. Role Guard - check if user is a system admin
    user_role = UserRole.objects.filter(user=request.user, role='sys_admin').first()
    if not user_role:
        return JsonResponse({"error": "Unauthorized"}, status=403) if request.headers.get('X-Requested-With') == 'XMLHttpRequest' else HttpResponse(status=403)

    sys_req = get_object_or_404(RequestedSystem, pk=pk)
    
    # 2. Scope Guard - verify system matches assigned system
    if sys_req.system != user_role.system_assigned:
        error_msg = "You cannot manage this system"
        return JsonResponse({"error": error_msg}, status=403) if request.headers.get('X-Requested-With') == 'XMLHttpRequest' else HttpResponse(status=403, content=error_msg)
    
    action = request.POST.get("action")
    comment = request.POST.get("comment", "")

    # 3. Validate Action
    if action not in ["approve", "reject"]:
        error_msg = "Invalid action"
        return JsonResponse({"error": error_msg}, status=400) if request.headers.get('X-Requested-With') == 'XMLHttpRequest' else redirect('system_admin_dashboard')

    # 4. Apply Decision
    if action == "approve":
        sys_req.sysadmin_status = "approved"
        sys_req.ict_status = "approved"
        decision_text = "Granted"
        badge_class = "bg-success"
    else:  # reject
        sys_req.sysadmin_status = "rejected"
        sys_req.ict_status = "rejected"
        decision_text = "Rejected"
        badge_class = "bg-danger"

    sys_req.sysadmin_comment = comment
    sys_req.sysadmin_decision_date = timezone.now()
    sys_req.system_admin = request.user
    sys_req.save()

    # 5. Notifications
    requester = sys_req.access_request.requester
    send_mail(
        subject=f"[TSC] Access Update for {sys_req.get_system_display()}",
        message=f"Dear {requester.full_name},\n\nRights have been granted/updated for {sys_req.get_system_display()}.\n\nRegards,\nTSC ICT Team",
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[requester.email],
        fail_silently=True,
    )

    # 6. Sync Parent Status
    sync_request_status(sys_req.access_request)

    # 7. Determine response type
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if is_ajax:
        # Return JSON for AJAX requests
        decision_date = sys_req.sysadmin_decision_date.strftime("%b %d, %Y %H:%M")
        return JsonResponse({
            "success": True,
            "message": f"Decision saved: {sys_req.get_system_display()} - {decision_text}",
            "system_id": sys_req.id,
            "status": decision_text,
            "badge_class": badge_class,
            "decision_date": decision_date,
            "comment": comment if comment else "-"
        })
    else:
        # Fallback to redirect for non-AJAX requests
        messages.success(request, "Decision saved.")
        return redirect(f"/access/system-admin/dashboard/?active_tab=pending&{request.GET.urlencode()}")





@login_required
def overall_admin_dashboard(request):
    # 1. Role Guard - User must be super admin or Overall Admin
    user_role = UserRole.objects.filter(user=request.user, role='super_admin').first()
    if not user_role and not request.user.is_superuser:
        messages.error(request, "You do not have access to Overall Admin dashboard.")
        return redirect('user_home')

    status_filter = request.GET.get("status", "")
    tsc_filter = request.GET.get("tsc", "")
    
    # ✅ NEW: Get Date Range
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    
    access_requests = AccessRequest.objects.all().select_related(
        'requester', 'directorate', 'hod_approver', 'ict_approver'
    ).prefetch_related('requested_systems').order_by('-submitted_at')

    # --- FILTERING ---
    if tsc_filter:
        access_requests = access_requests.filter(tsc_no__icontains=tsc_filter)
    
    if status_filter:
        access_requests = access_requests.filter(status=status_filter)

    # ✅ NEW: Date Range Filtering
    if start_date and end_date:
        try:
            s_date = datetime.strptime(start_date, "%Y-%m-%d")
            # Set end date to end of day (23:59:59)
            e_date = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            access_requests = access_requests.filter(submitted_at__range=(s_date, e_date))
        except ValueError:
            pass

    # --- EXPORT TO EXCEL ---
    if "export_excel" in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.title = "Access Requests"
        # Header
        ws.append(["Requester", "TSC No", "Directorate", "System", "HOD Status", "ICT Status", "SysAdmin Approver", "Submitted At"])
        
        for req in access_requests:
            for sys in req.requested_systems.all():
                sys_admin_name = sys.system_admin.full_name if sys.system_admin else "-"
                ws.append([
                    req.requester.full_name, req.tsc_no, 
                    req.directorate.name if req.directorate else "-",
                    sys.get_system_display(), sys.hod_status, sys.ict_status, sys_admin_name,
                    req.submitted_at.strftime("%Y-%m-%d %H:%M")
                ])
        
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="TSC_Requests_{localdate()}.xlsx"'
        wb.save(response)
        return response

    # --- EXPORT TO PDF ---
    if "export_pdf" in request.GET:
        buffer = io.BytesIO()
        # Landscape for better table fit
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        elements.append(Paragraph(f"TSC Access Report", styles["Title"]))
        if start_date and end_date:
             elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles["Normal"]))
        else:
             elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d')}", styles["Normal"]))
        
        elements.append(Spacer(1, 20))
        
        # Table Headers
        data = [["Requester", "TSC No", "System", "HOD", "ICT", "SysAdmin", "Date"]]
        
        for req in access_requests:
            for sys in req.requested_systems.all():
                sys_admin_name = sys.system_admin.full_name if sys.system_admin else "-"
                data.append([
                    req.requester.full_name, req.tsc_no, 
                    sys.get_system_display(), sys.hod_status.upper(), sys.ict_status.upper(), sys_admin_name,
                    req.submitted_at.strftime("%Y-%m-%d")
                ])
        
        # Styled Table
        table = Table(data, colWidths=[110, 60, 110, 60, 60, 90, 70])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')

    context = {
        "access_requests": access_requests,
        "total": RequestedSystem.objects.count(),
    }
    return render(request, "access_request/overall_admin_dashboard.html", context)



# In access_request/views.py

@require_POST
@login_required
def overall_admin_override(request, sys_id):
    if not request.user.is_superuser and getattr(request.user.userrole, 'role', '') != 'super_admin':
        return HttpResponse(status=403)

    system_request = get_object_or_404(RequestedSystem, id=sys_id)
    request_obj = system_request.access_request
    
    target_stage = request.POST.get('stage')
    new_status = request.POST.get('status')
    comment = request.POST.get('comment', f"Overridden by {request.user.full_name}")

    # ... (Keep your existing logic for HOD/ICT/SysAdmin updates) ...
    if target_stage == 'hod':
        system_request.hod_status = new_status
        system_request.hod_comment = comment
        system_request.hod_decision_date = timezone.now()
        request_obj.hod_approver = request.user 
        if new_status == 'rejected':
            system_request.ict_status = 'rejected'
            system_request.sysadmin_status = 'rejected'
    elif target_stage == 'ict':
        system_request.ict_status = new_status
        system_request.ict_comment = comment
        system_request.ict_decision_date = timezone.now()
        request_obj.ict_approver = request.user
    elif target_stage == 'sys_admin':
        system_request.sysadmin_status = new_status
        system_request.sysadmin_comment = comment
        system_request.system_admin = request.user
        system_request.sysadmin_decision_date = timezone.now()

    system_request.save()
    request_obj.save()
    sync_request_status(request_obj)

    # ✅ NEW: Send Notification Email
    send_mail(
        subject=f"[TSC] Admin Override: Access to {system_request.get_system_display()}",
        message=f"Dear {request_obj.requester.full_name},\n\n"
                f"Your request for {system_request.get_system_display()} has been updated by the System Administrator.\n\n"
                f"Stage: {target_stage.upper()}\n"
                f"New Status: {new_status.upper()}\n"
                f"Comment: {comment}\n\n"
                "Regards,\nTSC ICT Team",
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[request_obj.email],
        fail_silently=True,
    )

    messages.success(request, f"Override applied to {system_request.get_system_display()}. Email sent.")
    return redirect(f"/access/overall-admin/dashboard/?{request.GET.urlencode()}")

@login_required
def user_home(request):
    requests = AccessRequest.objects.filter(requester=request.user).prefetch_related("requested_systems").order_by('-submitted_at')
    return render(request, "access_request/user_home.html", {"requests": requests})

@login_required
def request_submitted(request): return render(request, 'access_request/request_submitted.html')

@login_required
def home_redirect(request):
    try: role = request.user.userrole.role
    except: return redirect("user_home")
    return redirect({"staff":"user_home","hod":"hod_dashboard","ict":"ict_dashboard","sys_admin":"system_admin_dashboard","super_admin":"overall_admin_dashboard"}.get(role,"user_home"))

def approve_request(request, request_id): return redirect('hod_dashboard')
def reject_request(request, request_id): return redirect('hod_dashboard')
@require_POST
def ict_approve(request, pk): return ict_system_decision(request, pk)
@require_POST
def ict_reject(request, pk): return ict_system_decision(request, pk)

@login_required
def export_system_admin_data(request, format):
    user_role = UserRole.objects.filter(user=request.user, role='sys_admin').first()
    if not user_role: 
        messages.error(request, "You are not assigned as a system admin."); 
        return redirect("home")
    requests = RequestedSystem.objects.filter(system_admin=request.user).select_related('access_request')
    if format == "csv":
        response = HttpResponse(content_type="text/csv")
        response['Content-Disposition'] = f'attachment; filename="system_admin_requests.csv"'
        writer = csv.writer(response)
        writer.writerow(["Requester", "TSC No", "System", "Request Type", "Access Level", "Status", "Date"])
        for r in requests: 
            writer.writerow([r.access_request.requester.full_name, r.access_request.tsc_no, r.get_system_display(), r.access_request.get_request_type_display(), r.level_of_access, r.sysadmin_status, r.access_request.submitted_at.strftime("%Y-%m-%d %H:%M")])
        return response
    return redirect("system_admin_dashboard")